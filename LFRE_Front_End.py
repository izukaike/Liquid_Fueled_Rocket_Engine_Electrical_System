'''
Contributors: Izuka Ikedionwu
    *As you edit,add, delete* please add your name :)*

Date Created: 10/12/23

Description: 
    GUI for the liquid fueled rocket engine. Interacts with Arduino through serial port.
    Using PyQtGraph for Interface with LabView Capabilities. Object Oriented with the main 
    class being RealTimePlotApp the initialization sets up all components or "widgets" followed
    by functions that control valves, igniter, and updating plots

Features:
    -  10 button fully functional graphical user interface 
    -  multi-threaded (read && write more efficiently)
    -  up to ~2ms graphic/response/record resolution (0.0019 sec) *recalc*
    -  asynchronous communication protocol
    -  two-factor verification
    -  automated data collection and data visualization
    -  Exception handling
    -  LabView extension
Work:
    - serial exception handling at the beginning of the program 
    - finish test procedure function in 2nd thread 
    - finish timing when it comes to pyserial commands (time.sleep(0.01) necessary?)
    - handling giant pressure spikes 
    - data analytics

Dependencies:
    pyqtgraph
    pyqt5
    numpy
    pyserial 
    subprocess
    openpyxl 
'''
import sys
import threading
import pyqtgraph as pg
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton,QHBoxLayout,QGridLayout,QSlider
from PyQt5.QtWidgets import QLabel
from PyQt5 import QtCore
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject, QMutex
from PyQt5 import uic
import numpy as np
import serial
import time 
from timeit import default_timer as timer
import subprocess
import numpy
from pyqtgraph.Qt import QtGui
import openpyxl
from openpyxl.chart import ScatterChart, Reference
import math
#---------------------------------------------------------------------------------------
#GLOBAL VARIABLES
#Worker Thread  Class
SPARKED = False
sLock  = threading.Lock()
class MyWorker(QObject):
    finished = pyqtSignal()
    #creates signal about thread function
    #test sequence function
    def test_sequence(self):
        global SPARKED
        global window
        #close both valves
        arduino.write(b'3')
        window.v1 = 0
        time.sleep(0.05)
        arduino.write(b'4')
        window.v2 = 0

        #wait 1.5 seconds then run sequence again
        time.sleep(1.5)
        #open valve 1
        arduino.reset_output_buffer()
        time.sleep(0.1)
        #open valve 1
        arduino.write(b'2')
        window.v2 = window.data_max
        time.sleep(0.001)
        #open valve 2
        arduino.write(b'1')
        window.v1 = window.data_max
        time.sleep(0.001)
        #start igniter
        start = time.time()
        sec10 = start
        i = 1

        while sec10 - start  < 5:
            arduino.write(b'7')
            window.i1 = window.data_max
            time.sleep(0.08)
            i  =  i +  1
            sec10 = time.time()
            if(SPARKED == True):
                break

        #close valve 1
        if(SPARKED == False):
            arduino.write(b'4')
            window.v1 = 0
            time.sleep(0.001)
            #close valve 2
            arduino.write(b'5')
            window.v2 = 0
            time.sleep(0.05)
            #end of sequence
            #arduino.reset_output_buffer()
        elif(SPARKED == True):
            time.sleep(1)
            arduino.write(b'4')
            window.v1 = 0
            time.sleep(0.001)
            #close valve 2
            arduino.write(b'5')
            window.v2 = 0
            time.sleep(0.05)
        
      
        SPARKED = False
        time.sleep(0.1)
        arduino.reset_output_buffer()
        self.finished.emit()
            
    #igniter function
    def igniter_sequence(self):
        arduino.reset_output_buffer()
        time.sleep(0.1)
        for i in range(5):
            arduino.write(b'7')
            time.sleep(0.08)
        #arduino.flush()
        time.sleep(0.1)
        arduino.reset_output_buffer()
        self.finished.emit()

#Thread class 
class MyThread(QThread):
    function_type = 0
    def run(self):
        self.worker = MyWorker()
        self.worker.finished.connect(self.quit)
        #chooses threaded function
        if self.function_type  == 1:
            self.worker.igniter_sequence()
        elif self.function_type == 2:
            self.worker.test_sequence()
#-------------------------------------------------------------------------------

#BGEINNING OF PROCESSING ( WHEN PROGRAM IS RAN)
#get com port // differs by cable
comport = input("Enter COMPORT:")

#Initialize Serial Communication with Arduino
#checks serial connection

arduino=serial.Serial() 
arduino.baudrate=115200#comm speed // has to be synced with arduino
arduino.port= comport
arduino.bytesize=serial.EIGHTBITS
#arduino.stopbits = serial.STOPBITS_ONE
#used when not reading data // changes when reading function
arduino.timeout = 0.01
arduino.write_timeout = 0




#opening excel file premature for less wait time in the future
dataframe = openpyxl.load_workbook("Recording_File.xlsx")
sheet = dataframe.active
sheet.delete_rows(2,sheet.max_row)
dataframe.save("C:\\Users\\izuka\\Documents\\A_New_Ard_Proj\\Recording_File.xlsx")
dataframe = openpyxl.load_workbook("Recording_File.xlsx")
sheet = dataframe.active

#opening texts file for real-time data collection in update function
txt_file = open('data.txt', 'w+')
txt_file.write("")
txt_file.close()
txt_file = open("data.txt", 'w')
if(txt_file):
    print("1: Raw File Ready 1/3")
else:
    print("1: Raw File Not Ready")
if(dataframe):
    print("2: Record File Available 2/3")
else:
    print("2: Record File Not Available")


#clears port buffer before anything
if arduino.is_open:
    arduino.close()
#try and catch exception handling here 
try:
    arduino.open()
except:
    i = 0
    while True and i < 3:
        arduino.port = comport
        time.sleep(0.5)
        arduino = serial.Serial()
        if arduino.is_open:
            break
        i += 1
    if( i == 3):
        print("Serial Connection Failed")
        sys.exit()

        

if arduino.is_open:
    print("3: Connected to Arduino 3/3")
else:
    print("3: Not Connected to Arduino")

#Needed delay for arduino to set up
time.sleep(2)

#do while to start GUI
start_sequence = input("Enter Start Code:")
while True:
    if(start_sequence == 'start'):
        break
    else:
        start_sequence = input("Enter Start Code:")


#Main UI Class
class LFRE_GUI_Control_App(QMainWindow):
    #Default Constructor
    def __init__(self):
        super().__init__()
        #class variables
        self.pause = 1
        #two factor logic
        self.arm1 = 0
        self.arm2 = 0
        self.armed = 0
        #plot logic
        self.starting_plot = 0
        #toggle logic
        self.is_on1 = False
        self.is_on2 = False
        self.is_on3 = True
        self.is_on4 = True
        self.is_on5 = True
        self.is_on6 = True
        self.is_on7 = True
        self.start_recording = 0
        self.i = 0
        self.j = 0
        self.v1 = 0
        self.start_bit = 0
        self.v2 = 0
        self.v3 = 0
        self.i1= 0
        self.test = 0
        self._finished_func = False
        self.buff_refresh = 1
        #data handling default values
        self.data_max = -100
        self.time = 0
        self.bytes_read = 24
        self.new_data1 = 0
        self.new_data2 = 0
        self.new_data3 = 0
        self.new_data4 = 0
        self.plot = 0
        self.domain = [1,2,3]
        self.status = [0,0,0,0,0,0,0,0]
        #thread opbject
        self.thread = [0,0]
        #plot buffer for opto
        self.str_data1 = '0' * 34
        print("type = " + str(type(self.str_data1)))
        self.str_data2 = "0" * 34
        self.str_data3 = "0" * 34
        self.str_data4 = "0" * 34
        self.label1 = {'font-size':'18px'}
        self.data_buff = [0,0,0,0]
        
        
        #FRONT END

        # Initialize the main window
        self.setWindowTitle("Liquid Fueled Rocket Engine GUI:The Linda")
        self.setGeometry(600, 150, 1200, 800)
        self.setStyleSheet("QMainWindow { background-color: light gray}")

        # Create a central widget to contain the plots and buttons
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Create layout for the central widget
        layout = QGridLayout(central_widget)
        # Create three PlotWidgets
        self.plot_widgets = [pg.PlotWidget() for _ in range(4)]
        
        #adding labels to graph
        self.plot_widgets[0].getPlotItem().setLabel('left',text = '<font color="black">Psi</font>')
        #self.plot_widgets[0].getPlotItem().setLabel('bottom', 'Time')
        self.plot_widgets[0].getPlotItem().setLabel('top', text = '<font color="black">Pressure Transducer 1</font>')


        self.plot_widgets[1].getPlotItem().setLabel('left',text = '<font color="black">Psi</font>')
        #self.plot_widgets[1].getPlotItem().setLabel('bottom', 'Time')
        self.plot_widgets[1].getPlotItem().setLabel('top','<font color="black">Pressure Transducer 2</font>')


        self.plot_widgets[2].getPlotItem().setLabel('left',text = '<font color="black">Psi<font>')
        #self.plot_widgets[2].getPlotItem().setLabel('bottom', 'Time')
        self.plot_widgets[2].getPlotItem().setLabel('top','<font color="black">Pressure Transducer 3</font>')

        self.plot_widgets[3].getPlotItem().setLabel('left',text = '<font color="black">Psi</font>')
        #self.plot_widgets[0].getPlotItem().setLabel('bottom', 'Time')
        self.plot_widgets[3].getPlotItem().setLabel('top', text = '<font color="black">Pressure Transducer 4</font>')

        self.bar_item = [0,1,3,5]
        self.bar_item[0] = pg.BarGraphItem(x=[1], height=60, width=1, brush=(169,169,169))
        self.plot_widgets[0].addItem(self.bar_item[0])
        self.plot_widgets[0].setBackground('white')

        self.bar_item[1] = pg.BarGraphItem(x=[1], height=33, width=1, brush=(169,169,169))
        self.plot_widgets[1].addItem(self.bar_item[1])
        self.plot_widgets[1].setBackground('white')

        self.bar_item[2] = pg.BarGraphItem(x=[1], height=85, width=1, brush=(169,169,169))
        self.plot_widgets[2].addItem(self.bar_item[2])
        self.plot_widgets[2].setBackground('white')

        self.bar_item[3] = pg.BarGraphItem(x=[1], height=60, width=1, brush=(169,169,169))
        self.plot_widgets[3].addItem(self.bar_item[3])
        self.plot_widgets[3].setBackground('white')
       
        #adding bar to graphs with y range max 
        for i in range(4):
            self.plot_widgets[i].setYRange(0, 500, padding=0)
        
        #adding widgets to GUI
        number_of_plots = 4
        for i in range(number_of_plots):
            layout.addWidget(self.plot_widgets[i],0,i)

        #adding text widgets
        self.label1 = QLabel("  Psi = -- Time = -- RR = --")
        self.label2 = QLabel("  Psi = -- Time = -- RR = --")
        self.label3 = QLabel("  Psi = -- Time = -- RR = --")
        self.label4 = QLabel("  Psi = -- Time = -- RR = --")
        self.label1.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        self.label2.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        self.label3.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        self.label4.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        
        layout.addWidget(self.label1)
        layout.addWidget(self.label2)
        layout.addWidget(self.label3)
        layout.addWidget(self.label4)

        
        

        # Create toggle buttons
        self.switch1 = QPushButton("VALVE 1: CLOSED")
        self.switch2 = QPushButton("VALVE 2: CLOSED")
        self.switch3 = QPushButton("VALVE 3: CLOSED")
        self.switch4 = QPushButton("IGNITER")
        self.switch5 = QPushButton("ARMED1: OFF")
        self.switch6 = QPushButton("ARMED2: OFF")
        self.switch7 = QPushButton("RECORDING: OFF")

        #switch 1 setup
        self.switch1.setCheckable(True)
        self.switch1.clicked.connect(self.toggle_switch1)
        self.switch1.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")
        #layout.addWidget(self.switch1)

        #switch 2 setup
        self.switch2.setCheckable(True)
        self.switch2.clicked.connect(self.toggle_switch2)
        self.switch2.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")
        #layout.addWidget(self.switch2)

        #switch 3 setup
        #self.switch3.setCheckable(True)
        #self.switch3.clicked.connect(self.toggle_switch3)
        #self.switch3.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")
        #layout.addWidget(self.switch3)

        self.switch4.setCheckable(True)
        self.switch4.clicked.connect(self.toggle_switch4)
        self.switch4.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 20px; }")
        #layout.addWidget(self.switch4)

        #switch 5 setup
        self.switch5.setCheckable(True)
        self.switch5.clicked.connect(self.toggle_switch5)
        self.switch5.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
        #layout.addWidget(self.switch5)

        #switch 6 setup
        self.switch6.setCheckable(True)
        self.switch6.clicked.connect(self.toggle_switch6)
        self.switch6.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
        #layout.addWidget(self.switch6)

        self.switch7.setCheckable(True)
        self.switch7.clicked.connect(self.toggle_switch7)
        self.switch7.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
        #layout.addWidget(self.switch7)

        layout.addWidget(self.switch1)
        layout.addWidget(self.switch2)
        layout.addWidget(self.switch4)
        layout.addWidget(self.switch7)
        layout.addWidget(self.switch5)
        layout.addWidget(self.switch6)


        number_button = 5
        #initializes 2 buttons
        self.buttons = [QPushButton(f"Valve {i+1}") for i in range(number_button)]
        for i, button in enumerate(self.buttons):
            
            button.setStyleSheet("QPushButton { background-color:white; color: black; font-size: 20px; }")
            #button setup 
            j = 0
            if(i == 0):
                button.setText("START PLOT")
                button.clicked.connect(lambda: self.start_plot())  # Connect button click event
            if(i == 3):
                button.setText("ABORT")
                button.clicked.connect(lambda: self.abort())  # Connect button click event
                button.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
            if(i == 2):
                button.setText("LabView(Windows)")
                button.clicked.connect(lambda: self.runLabView_Windows())  # Connect button click event
            if(i == 1):
                button.setText("TEST PROCEDURE: NOT ARMED")
                button.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
                button.clicked.connect(lambda: self.start_test())  # Connect button click event
            if(i == 4):
                button.setText("CALIBRATE")
                button.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 18px; }")
                button.clicked.connect(lambda: self.start_calibration())  # Connect button click event
            
        layout.addWidget(self.buttons[1])
        for i in range(number_button):
            #if( i == 4):
            if i != 1:
                layout.addWidget(self.buttons[i])

        # Initialize data for the plots
        number_of_points = 4
        self.data = [np.zeros(number_of_points) for _ in range(4)]

        # Initialize the plots
        self.curves = [pw.plot(self.data[i]) for i, pw in enumerate(self.plot_widgets)]

        #set x limit and y limit
        for pw in self.plot_widgets:
            pw.setRange(xRange=[0,2], yRange=[0, 1000])
        
        # Start a timer to update the plots at regular intervals 
        refresh = 0 # 1ms
        self.timer = pg.QtCore.QTimer(self)
        self.timer.setInterval(refresh)
        self.timer.timeout.connect(self.update_plots)
        self.timer.start(refresh)  # Update every 100 ms

#end of setup code
#----------------------------------------------------------------------------------
#Class Functions
    def toggle_switch1(self):
        sender = self.sender()  # Get the button that emitted the signal
        sender.isChecked = 0
        #NEW CODE
        self.is_on1 = not self.is_on1
        if self.is_on1:
            self.v1 = 0
            self.control_valve1_open()
            self.switch1.setText("VALVE1: CLOSED")
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
        else:
            self.v1 = self.data_max
            self.control_valve1_close()
            sender.setText("VALVE 1:  OPEN")
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        '''
        if sender.isChecked():
            #important line that calls valve opening/closing function the rest is aesthetic
            self.control_valve1_open()
            self.setText("VALVE1:CLOSED")
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
        else:
            self.control_valve1_close()
            sender.setText("VALVE 1:  OPEN")
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        '''
    
    def toggle_switch2(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on2 = not self.is_on2
        if self.is_on2:
            self.v2 = 0
            self.control_valve2_open()
            self.switch2.setText("VALVE2: CLOSED")
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
        else:
            self.v2 = self.data_max
            self.control_valve2_close()
            sender.setText("VALVE 2:  OPEN")
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

    def toggle_switch4(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on4 = not self.is_on4
        if self.is_on4:
            self.i1 = self.data_max
            # turns on both
            if self.is_on1 == False and self.is_on2 == False:
                self.control_igniter_open()
                self.switch4.setText("IGNITER")
                sender.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 18px; }")
        else:
            if self.is_on1 == False and self.is_on2 == False:
                self.i1 = self.data_max
                self.control_igniter_open()
                self.switch4.setText("IGNITER")
                sender.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 18px; }")

    def toggle_switch5(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on5 = not self.is_on5
        if self.is_on5:
            self.switch5.setText("ARMED1: OFF")
            self.arm1 = 0
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
            if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        else:
            self.switch5.setText("ARMED1: ON")
            self.arm1 = 1
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
    
    def toggle_switch6(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on6 = not self.is_on6
        if self.is_on6:
            self.switch6.setText("ARMED2: OFF")
            self.arm2 = 0
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
            if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        else:
            self.switch6.setText("ARMED2: ON")
            self.arm2= 1
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            if self.arm1  == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

    def toggle_switch7(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on7= not self.is_on7
        if self.is_on7:
            self.start_recording = 0
            self.switch7.setText("RECORDING: OFF")
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        else:
            #add recording function 
            self.start_recording = 1
            self.switch7.setText("RECORDING: ON")
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
    
    #function that communicates with Arduino and opens valve1
    def control_valve1_open(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'1')
        #time.sleep(0.1)
    
    def control_valve1_close(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'4')
        #arduino.flush()
        #time.sleep(0.1)

    def control_valve2_open(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'2')
        #arduino.flush()
        #time.sleep(0.1)
        #arduino.close()

    def control_valve2_close(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'5')
        #arduino.flush()
        #time.sleep(0.1)

    def control_valve3_open(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'3')
        #arduino.flush()
        #time.sleep(0.1)

    def control_valve3_close(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'6')
        #arduino.flush()
        #time.sleep(0.1)

# 1 Threaded function
    def control_igniter_open(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        self.thread[0] = MyThread() # added function number
        self.thread[0].function_type = 1
        self.thread[0].start()
        self.thread[0].finished.connect(self.on_finished1)
        #time.sleep(0.1)
        
    def on_finished1(self):
        #self.thread[0].wait()
        print("Igniter Thread is Finished, Bytes in waiting = " + str(arduino.in_waiting))

   # 2 Threaded function
    def start_test(self):
        global comport
        global arduino
        global i 

        #opens valve on gui
        #starts test sequence thread
        if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
            self.is_on1 = True
            self.is_on2 = True
            self.switch1.setText("VALVE1: OPEN")
            self.switch1.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
            self.switch2.setText("VALVE2: OPEN")
            self.switch2.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")

            arduino.write(b'!')
            self.buttons[1].setText("***TEST PROCEDURE")
            self.thread[1] = MyThread()
            self.thread[1].function_type = 2
            self.thread[1].start()
            self.thread[1].finished.connect(self.on_finished2)


    def on_finished2(self):
        print("Test Thread Finished")
        self.buttons[1].setText("TEST PROCEDURE: ARMED")

        self.is_on1 = False
        self.is_on2 = False
        self.switch1.setText("VALVE1: CLOSE")
        self.switch1.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")
        self.switch2.setText("VALVE2: CLOSE")
        self.switch2.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")

    def start_calibration(self):
        #calibrates data no need for multithreading
        arduino.write(b'@')
        time.sleep(4)# wait 4 seconds
        self.buttons[0].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
        self.test = 1
        if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
            self.buttons[1].setText("TEST PROCEDURE: ARMED")
            self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
        else:
            self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
            self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

        #arduino.flush()

    def runLabView_Windows(self):
        #this is function is operating system specific and pretty much goes into the command line and oepns the LabView app
        subprocess.run('cd C:\\Program Files (x86)\\National Instruments\\LabVIEW 2023 & dir & start LabVIEW', shell = True)
    
    def start_plot(self):
        if(self.test == 1):
            self.starting_plot = 1
            #arduino.reset_input_buffer()
            print("start")
            #print(arduino.in_waiting)
            arduino.write(b'9')
            time.sleep(0.05)
            arduino.write(b'(')
            time.sleep(0.05)
            arduino.write(b'^')
            #arduino.flush()
            #increase timeout time to avoid write timeout exception 
            arduino.timeout = 1
            #start time for data analysis
            self.time = time.time()
            # 0.1 delay from start to graph for more accurate reading
            #time.sleep(0.1)

#Still needs to finish functionality
    def abort(self):
        #close valve 1
        global SPARKED
        SPARKED = True
        arduino.write(b'4')
        time.sleep(0.05)
        #close valve 2
        arduino.write(b'5')
        #print("working")

    #main function that updates plots
    def update_plots(self):
        begin = time.time()
        bits1 = b''
        #status = b''
        
        #t = time.time()
        if self._finished_func == False and arduino.in_waiting > 0:
            #reads status flag
            temp = arduino.read()
            #print(temp)
            if temp != b'' and len(temp) == 1 and temp != b'\n' and temp != b'\r':
                temp = temp[0]
                
                for i in range(8):
                    if( (temp & 1) == 1):
                        self.status[i] = 1
                    else:
                        self.status[i] = 0
                    temp = temp >> 1
            
            #make sure on start up there are is no 1 in 1 place for all incoming data
            if(self.status == [1,0,0,0,0,0,0,0]):
                self.start_bit = 1
                

            if(self.start_bit == 1 and self.status[0] == 1):
                #print(self.status)
                bits1 = arduino.read(self.bytes_read)
        
        if(len(bits1) == self.bytes_read and bits1 != '' and self.starting_plot == 1 and self._finished_func == False):
            
            rounds = [0,8,16]
            offset  = [0,2,4,6]
            pressures = [0] * 12
            index = 0
            #print(bits1)
            for i in range(4):
            #round is 3 offset is 4
                for j in range(3):
                    pt_low = bits1[rounds[j]*1 + offset[i] ]
                    pt_high = bits1[rounds[j]*1 + (offset[i]+1)]
                    #print("low = " + str(pt_low))
                    #print("high = " + str(pt_high))
                    pressures[index] = (pt_high << 8) | pt_low
                    index += 1

            arr = [0] * 3
            medians  = []
            for i in range(1,len(pressures)+1):
                arr[(i) % 3] = pressures[i-1]
                if i % 3 == 0:
                    arr.sort()
                    medians.append(arr[1])
            
            #print(medians)
            if(medians[0] < 1023 or medians[0] >= 0  ):
                self.new_data1 = medians[0]
            else:
                print(medians[0])
        
            self.bar_item[0].setOpts(x=self.domain, height=self.new_data1)
            #checks pressure for ignition in  seperate thread
            if(self.new_data1 > 0): 
                SPARKED = True
            
            if(medians[1] < 1023 or medians[1] >= 0  ):
                self.new_data2 = medians[1]
            else:
                print(medians[1])
            self.bar_item[1].setOpts(x=self.domain, height=self.new_data2)

            if( medians[2] < 1023 or medians[2] >= 0 ):
                self.new_data3 = medians[2]
            else:
                print(medians[2])
            self.bar_item[2].setOpts(x=self.domain, height=self.new_data3)

            if(medians[3] < 1023 or medians[3] > 0  ):
                self.new_data4 = medians[3]
            else:
                print(medians[3])
            self.bar_item[3].setOpts(x=self.domain, height=self.new_data4)
            
            if( self.start_recording == 1):
               txt_file.write(str(self.new_data1)+"\n"+str(self.new_data2)+"\n"+str(self.new_data3)+"\n"+str(self.new_data4)+"\n"+str(self.status[6]*-100)+"\n")  
               txt_file.write(str(self.status[5]*-90)+"\n"+str(self.status[4]*-80)+"\n"+str(time.time()- self.time)+"\n") #str(self.status[0]*-100
               txt_file.write(str(self.status[3]*-175)+"\n"+str(self.status[2]*-170)+"\n" +str(self.status[1]*-165)+"\n" + str(self.status[7]*-160)+"\n") 
               self.i += 1
                
        #calls abort function
        '''
        if(self.status[7] == 1 or self.status[1] == 1 or self.status[2] == 1 or self.status[3] == 1):
            self.abort()
        '''
        
        #q1 = time.time()
        refresh_rate = str(math.ceil((time.time() - begin) * 1_000))
        curr_time = round( ((time.time() - self.time)),2)
        if curr_time > 100:
           curr_time *= round(pow(10,-9),2)
        str_data1 = "  Psi = " + str(self.new_data1) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        str_data2 = "  Psi = " + str(self.new_data2) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        str_data3 = "  Psi = " + str(self.new_data3) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        str_data4 = "  Psi = " + str(self.new_data4) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        #print(str(self.new_data1))
            
        self.label1.setText(str_data1)
        self.label2.setText(str_data2)
        self.label3.setText(str_data3)
        self.label4.setText(str_data4)
        #print("process = " +  str((time.time()- t)*1_000))
        
        
#---------------------Main Function-----------------------------------------
if __name__ == '__main__':
    #Application set up
    app = QApplication(sys.argv)
    window = LFRE_GUI_Control_App()
    arduino.reset_input_buffer()
    window.show()
    app.exec_()
    txt_file.close()
    #recording data from text file into excel
    txt_file = open("data.txt",'r')
    count = 0
    for i in range(window.i):
        for j in range(12):
            rec_data1 = sheet.cell(row = i+2, column = j+1)
            rec_data1.value = str("=" + txt_file.readline()) 
        count += 1

    dataframe.save("C:\\Users\\izuka\\Documents\\A_New_Ard_Proj\\Recording_File.xlsx")
    print("Data Is Saved")
    if window.start_recording == 1:
        subprocess.run('C:\\Users\\izuka\Documents\\A_New_Ard_Proj\\Recording_File.xlsx', shell = True)
    #clear all buffers when something messes up
    arduino.reset_output_buffer()
    arduino.close()
    sys.exit()
    print(i)
