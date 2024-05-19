'''
Contributors: Izuka Ikedionwu
    *As you edit,add, delete* please add your name :)*

Date Created: 10/12/23

Description: 
    GUI for the liquid fueled rocket engine. Interacts with Arduino through serial port.
    Using PyQtGraph for Interface with LabView Capabilities. Object Oriented with the main 
    class being lfre gui the initialization sets up all components or "widgets" followed
    by functions that control valves, igniter, and updating plots

Features:
    -  10 button fully functional graphical user interface 
    -  multi-threaded (read && write more efficiently for parallel DAQ
    -  > 2ms the app reads 24 bytes at 115200 bits/s
    -  asynchronous communication protocol
    -  advanced filtering techniques
    -  two-factor verification
    -  automated data collection and data visualization
    -  Exception handling
    -  LabView extension

Work:
    - serial exception handling at the beginning of the program 
    - finish test procedure function in 2nd thread 
    - finish timing when it comes to pyserial commands (time.sleep(0.01) necessary?)
    - handling giant pressure spikes 
    - data analytics

Dependencies:
    pyqtgraph
    pyqt5
    numpy
    pyserial 
    subprocess
    openpyxl 
'''
import sys
import threading
import pyqtgraph as pg
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton,QHBoxLayout,QGridLayout,QSlider
from PyQt5.QtWidgets import QLabel
from PyQt5 import QtCore
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject, QMutex
from PyQt5 import uic
import numpy as np
import serial
import time 
from timeit import default_timer as timer
import subprocess
import numpy
from pyqtgraph.Qt import QtGui
import openpyxl
from openpyxl.chart import ScatterChart, Reference
import math
import tkinter as tk
from tkinter import ttk
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from random import randint

'''
Classes:
    MyWorker: worker thread
    Thread: main thread class
    LFRE_GUI: Main GUI
'''
#---------------------------------------------------------------------------------------
#GLOBAL VARIABLES
SPARKED = False
'''
   thread worker class
'''
class MyWorker(QObject):
    # creates signal uses to comm with main thread
    finished = pyqtSignal()
    '''
        sends test sequence to microcontroller through serial comm

        passes obj

        returns nothing 
    '''
    def test_sequence(self):

        #access sparked variables *** this should be a member variable in gui
        global SPARKED
        global window
        global app


        time.sleep(1)
        #close both valves
        window.control_valve1_close()
        time.sleep(0.001)
        window.control_valve2_close()

        window.v1 = 0
        window.v2 = 0

        #START PHASE 1
        try: 
            #wait 0.5 seconds then run sequence again
            time.sleep(0.5)

            #reset output so test sequence is first in buff
            arduino.reset_output_buffer()

            #waits for serial comm to refresh on mcu side
            time.sleep(0.5)
            
            '''
            I wanted to send the command straight from the serial port
            and not the gui obj to try and eliminate the obj being used in 2 threads 
            with out more robusts safety code 
            '''
            #open valve 1
            window.control_valve1_open()
            window.v2 = window.data_max

            time.sleep(0.001)

            #open valve 2
            window.control_valve2_open()
            window.v1 = window.data_max

            time.sleep(0.001)
        except:
            print("phase 1 of test failed")

        #START PHASE 2
        try: 

            start = time.time()
            five_seconds = start
            i = 1
            
            #run sparks for 5 seconds
            while five_seconds - start  < 5:

                #spark coil 
                window.control_igniter_open()
                window.i1 = window.data_max
        
                #sparks every 80ms
                '''
                CHANGE THIS LINE TO CHANGE SPARK SPEED
                '''
                time.sleep(0.08)
                i  =  i +  1

                five_seconds = time.time()

                #checks with gui if combustion is achieved changed in update_gui function
                if(SPARKED == True):
                    break
        except:
            print("phase 2 of test failed")

        #START PHASE 3
        try:
            #if after 5 seconds no combustion then...
            if(SPARKED == False):
                
                #close valve 1
                window.control_valve1_close()
                window.v1 = 0
                time.sleep(0.001)

                #close valve 2
                window.control_valve2_close()
                window.v2 = 0
                time.sleep(0.05)
            #if combustion achieved
            elif(SPARKED == True):
                #keep valves open for 1 second
                time.sleep(1)

                #close both valves
                window.control_valve1_close()
                window.v1 = 0

                time.sleep(0.001)
                #close valve 2
                window.control_valve2_close()
                window.v2 = 0

                time.sleep(0.05)
        except:
            print("phase 3 of test failed")
        
        #resets sparked var for another test
        SPARKED = False
        time.sleep(0.1)
        #may not really be needed 
        arduino.reset_output_buffer()
        time.sleep(0.5)
        #sends finished signal to main thread
        self.finished.emit()
            
    '''
    sparks only igniter

    passes obj

    returns nothing
    '''
    def igniter_sequence(self):
        #resets output buffer so coil signal is first
        #sparks coil 5 times every 80 ms
        arduino.write(b'7')
        self.finished.emit()

'''
main thread class
'''
class MyThread(QThread):
    function_type = 0
    '''
    automatically runs when api start function
    which runs this function

    passes obj

    returns nothing
    '''
    def run(self):
        #calles worker constructor
        self.worker = MyWorker()

        #quits thread after run function finished
        self.worker.finished.connect(self.quit)

        #chooses threaded function
        if self.function_type  == 1:

            self.worker.igniter_sequence()

        elif self.function_type == 2:

            self.worker.test_sequence()

#-------------------------------------------------------------------------------

#Main UI Class
class LFRE_GUI_Control_App(QMainWindow):
    #Default Constructor
    def __init__(self):
        #inheritance
        super().__init__()
        #class variables
        self.test_done = ''
        self.test_start = ''

        #two factor logic
        self.arm1 = 0
        self.arm2 = 0
        self.armed = 0

        #plot logic
        self.starting_plot = 0

        #toggle logic
        self.is_on1 = False
        self.is_on2 = False
        self.is_on3 = True
        self.is_on4 = True
        self.is_on5 = True
        self.is_on6 = True
        self.is_on7 = True

        #button logic
        self.start_recording = 0

        #device recording
        self.i = 0
        self.j = 0
        self.v1 = 0
        self.v2 = 0
        self.i1= 0

        #data sync logic
        self.start_bit = 0
        self.test = 0

        #data handling default values
        self.data_max = 1
        self.time = 0
        self.bytes_read = 24
        self.new_data1 = 0
        self.new_data2 = 0
        self.new_data3 = 0
        self.new_data4 = 0
        self.plot = 0
        self.domain = [1,2,3]
        self.status = [0,0,0,0,0,0,0,0]
        #thread opbject
        self.thread = [0]*2

        #plot string buffer for labels 
        self.str_data1 = '0' * 34
        self.str_data2 = "0" * 34
        self.str_data3 = "0" * 34
        self.str_data4 = "0" * 34
        self.label1 = {'font-size':'18px'}
        self.data_buff = [0,0,0,0]

        #initializes recording data for main loop speed
        self.recording_data = []
        
        #FRONT END SETUP

        #start here

        # Initialize the main window
        #top left
        self.setWindowTitle("Liquid Fueled Rocket Engine GUI:The Linda")
        #sets mini window left of screen for control and view of terminal
        self.setGeometry(600, 150, 1200, 800)
        self.setStyleSheet("QMainWindow { background-color: light gray}")

        # Create a central widget to contain the plots and buttons
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # Create layout for the central widget
        layout = QGridLayout(central_widget)
        # Create 4 PlotWidgets                       *magic number*
        self.plot_widgets = [pg.PlotWidget() for _ in range(5)]
        
        #adding labels to graph
        # no loops for change in title if possible make this into a loop
        # text in html 
        self.plot_widgets[0].getPlotItem().setLabel('left',text = '<font color="black">Psi</font>')
        self.plot_widgets[0].getPlotItem().setLabel('top', text = '<font color="black">Pressure Transducer 1</font>')

        self.plot_widgets[1].getPlotItem().setLabel('left',text = '<font color="black">Psi</font>')
        self.plot_widgets[1].getPlotItem().setLabel('top','<font color="black">Pressure Transducer 2</font>')

        self.plot_widgets[2].getPlotItem().setLabel('left',text = '<font color="black">Psi<font>')
        self.plot_widgets[2].getPlotItem().setLabel('top','<font color="black">Pressure Transducer 3</font>')

        self.plot_widgets[3].getPlotItem().setLabel('left',text = '<font color="black">Psi</font>')
        self.plot_widgets[3].getPlotItem().setLabel('top', text = '<font color="black">Pressure Transducer 4</font>')

        self.plot_widgets[4].getPlotItem().setLabel('left',text = '<font color="black">Psi</font>')
        self.plot_widgets[4].getPlotItem().setLabel('top', text = '<font color="black">Plot</font>')

        #creates 4 bar items adds them to gui and gives init value
        self.bar_item = [0]*4
        self.bar_item[0] = pg.BarGraphItem(x=[1], height=60*5, width=1, brush=(169,169,169))
        self.plot_widgets[0].addItem(self.bar_item[0])
        self.plot_widgets[0].setBackground('white')

        self.bar_item[1] = pg.BarGraphItem(x=[1], height=33*5, width=1, brush=(169,169,169))
        self.plot_widgets[1].addItem(self.bar_item[1])
        self.plot_widgets[1].setBackground('white')

        self.bar_item[2] = pg.BarGraphItem(x=[1], height=85*5, width=1, brush=(169,169,169))
        self.plot_widgets[2].addItem(self.bar_item[2])
        self.plot_widgets[2].setBackground('white')

        self.bar_item[3] = pg.BarGraphItem(x=[1], height=60*5, width=1, brush=(169,169,169))
        self.plot_widgets[3].addItem(self.bar_item[3])
        self.plot_widgets[3].setBackground('white')

        #self.plot_widgets[4].setBackground('white')
       
        #adding bar to graphs with y range max 
        for i in range(4):                        #35 psi le way for looks only
            self.plot_widgets[i].setYRange(0, 1000, padding=25)
        
        #adding widgets to GUI
        number_of_plots = 4
        for i in range(number_of_plots):
            layout.addWidget(self.plot_widgets[i],0,i)
        #layout.addWidget(self.plot_widgets[4],0,4,4,6)

        #adding text widgets fpr ducer and gui data
        self.label1 = QLabel("  Psi = -- Time = -- RR = --")
        self.label2 = QLabel("  Psi = -- Time = -- RR = --")
        self.label3 = QLabel("  Psi = -- Time = -- RR = --")
        self.label4 = QLabel("  Psi = -- Time = -- RR = --")
        self.label1.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        self.label2.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        self.label3.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        self.label4.setStyleSheet("QLabel { background-color: white; color: black; font-size: 20.5px; }")
        
        #adding to gui
        layout.addWidget(self.label1)
        layout.addWidget(self.label2)
        layout.addWidget(self.label3)
        layout.addWidget(self.label4)

        # Create toggle buttons
        self.switch1 = QPushButton("VALVE 1: CLOSED")
        self.switch2 = QPushButton("VALVE 2: CLOSED")
        self.switch3 = QPushButton("VALVE 3: CLOSED")
        self.switch4 = QPushButton("IGNITER")
        self.switch5 = QPushButton("ARMED1: OFF")
        self.switch6 = QPushButton("ARMED2: OFF")
        self.switch7 = QPushButton("RECORDING: OFF")

        #switch 1 setup
        #make a for loop for this 
        self.switch1.setCheckable(True)
        self.switch1.clicked.connect(self.toggle_switch1)
        self.switch1.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")

        #switch 2 setup
        self.switch2.setCheckable(True)
        self.switch2.clicked.connect(self.toggle_switch2)
        self.switch2.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")

        self.switch4.setCheckable(True)
        self.switch4.clicked.connect(self.toggle_switch4)
        self.switch4.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 20px; }")

        #switch 5 setup
        self.switch5.setCheckable(True)
        self.switch5.clicked.connect(self.toggle_switch5)
        self.switch5.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")

        #switch 6 setup
        self.switch6.setCheckable(True)
        self.switch6.clicked.connect(self.toggle_switch6)
        self.switch6.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")

        self.switch7.setCheckable(True)
        self.switch7.clicked.connect(self.toggle_switch7)
        self.switch7.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
       
        #adding to gui 
        layout.addWidget(self.switch1)
        layout.addWidget(self.switch2)
        layout.addWidget(self.switch4)
        layout.addWidget(self.switch7)
        layout.addWidget(self.switch5)
        layout.addWidget(self.switch6)

        #adding buttons to gui 
        number_button = 6
        #may use this syntax to loop texts
        self.buttons = [QPushButton(f"Valve {i+1}") for i in range(number_button)]
        for i, button in enumerate(self.buttons):
            #default for all buttons
            button.setStyleSheet("QPushButton { background-color:white; color: black; font-size: 20px; }")
            
            #button setup 
            j = 0

            #change these numbers to change button order
            if(i == 0):
                button.setText("START PLOT")           
                button.clicked.connect(lambda: self.start_plot())  # Connect button click event
            if(i == 3):
                button.setText("ABORT")
                button.clicked.connect(lambda: self.abort())  # Connect button click event
                button.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
            if(i == 2):
                button.setText("LabView(Windows)")
                button.clicked.connect(lambda: self.runLabView_Windows())  # Connect button click event
            if(i == 1):
                button.setText("TEST PROCEDURE: NOT ARMED")
                button.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
                button.clicked.connect(lambda: self.start_test())  # Connect button click event
            if(i == 4):
                button.setText("CALIBRATE")
                button.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 18px; }")
                button.clicked.connect(lambda: self.start_calibration())  # Connect button click event
            if(i == 5):
                button.setText("PASS/FAIL")
                button.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 18px; }")
                button.clicked.connect(lambda: print("ADD FUNCTIONALITY"))  # Connect button click event
            
        #adding to gui
        layout.addWidget(self.buttons[1])
        for i in range(number_button):
            #if( i == 4):
            if i != 1:
                layout.addWidget(self.buttons[i])

        # Initialize data for the plots
        number_of_points = 5
        self.data = [np.zeros(number_of_points) for _ in range(5)]

        # Initialize the plots
        self.curves = [pw.plot(self.data[i]) for i, pw in enumerate(self.plot_widgets)]

        #set x limit and y limit
        for pw in self.plot_widgets:
            pw.setRange(xRange=[0,2], yRange=[0, 1000])
        
        # Start a timer to update the plots at regular intervals 
        refresh = 0 
        self.timer = pg.QtCore.QTimer(self)
        self.timer.setInterval(refresh)
        self.timer.timeout.connect(self.update_plots)
        self.timer.start(refresh)  

#end of setup code
#----------------------------------------------------------------------------------
#Class Functions
    ''' switch buttonfunction for valve 1

        passes obj

        returns nothing
    '''
    def toggle_switch1(self):
        #learn more about this arch
        sender = self.sender()  # Get the button that emitted the signal
        sender.isChecked = 0
        
        #toggle logic
        self.is_on1 = not self.is_on1
        if self.is_on1:
            self.v1 = 0
            self.control_valve1_open()
            self.switch1.setText("VALVE1: OPEN")
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        else:
            self.v1 = self.data_max
            self.control_valve1_close()
            sender.setText("VALVE 1: CLOSE")
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")

    ''' switch button function for valve 2

        passes obj

        returns nothing
    '''
    def toggle_switch2(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on2 = not self.is_on2
        if self.is_on2:
            self.v2 = 0
            self.control_valve2_open()
            self.switch2.setText("VALVE2: OPEN")
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        else:
            self.v2 = self.data_max
            self.control_valve2_close()
            sender.setText("VALVE 2: CLOSE")
            sender.setStyleSheet("QPushButton { background-color:green; color: black; font-size: 18px; }")

    ''' switch buttonfunction for coil

        passes obj

        returns nothing
    '''
    def toggle_switch4(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on4 = not self.is_on4
        if self.is_on4:
            self.i1 = self.data_max
            # turns on both
            if self.is_on1 == False and self.is_on2 == False:
                self.control_igniter_open()
                self.switch4.setText("IGNITER")
                sender.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 18px; }")
        else:
            if self.is_on1 == False and self.is_on2 == False:
                self.i1 = self.data_max
                self.control_igniter_open()
                self.switch4.setText("IGNITER")
                sender.setStyleSheet("QPushButton { background-color: white; color: black; font-size: 18px; }")

    ''' switch for 2 step test verification

        passes obj

        returns nothing
    '''
    def toggle_switch5(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on5 = not self.is_on5
        if self.is_on5:
            self.switch5.setText("ARMED1: OFF")
            self.arm1 = 0
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

            #test button logic
            if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

        else:
            self.switch5.setText("ARMED1: ON")
            self.arm1 = 1
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            
            #test button logic
            if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

    '''
    switch for 2step test procedure

    passes obj 

    return nothing
    '''
    def toggle_switch6(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on6 = not self.is_on6
        if self.is_on6:
            self.switch6.setText("ARMED2: OFF")
            self.arm2 = 0
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
            
            #test button logic
            if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        
        else:
            self.switch6.setText("ARMED2: ON")
            self.arm2= 1
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            
            #test button logic
            if self.arm1  == 1 and self.arm2 == 1 and self.test == 1:
                self.buttons[1].setText("TEST PROCEDURE: ARMED")
                self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
            else:
                self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")            
                self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

    '''
    recording switch button

    passes obj

    return nothing
    '''
    def toggle_switch7(self):
        sender = self.sender()  # Get the button that emitted the signal
        self.is_on7= not self.is_on7
        if self.is_on7:
            self.start_recording = 0
            self.switch7.setText("RECORDING: OFF")
            sender.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")
        else:
            
            self.start_recording = 1
            self.switch7.setText("RECORDING: ON")
            sender.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
    
    '''
    function that communicates with Arduino and opens valve1

    passes obj

    return nothing
    '''
    def control_valve1_open(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'1')
    
    '''
    closes valve 1

    passes obj 

    returns nothing
    '''
    def control_valve1_close(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'4')

    '''
    open valve 1

    passes obj 

    returns nothing
    '''
    def control_valve2_open(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'2')

    '''
    closes valve 2

    passes obj 

    returns nothing
    '''
    def control_valve2_close(self):
        #valve 1 sends 1 to arduino and arduino reads 1 and knows to activate valve1
        arduino.write(b'5')

    '''
    activates coil in seperate thread\
    
    passes obj

    returns nothing
    '''
    def control_igniter_open(self):
        arduino.write(b'7')
        '''
        #instantiates thread obj 
        self.thread[0] = MyThread() # added function number
        self.thread[0].function_type = 1
        #starts run function
        self.thread[0].start()
        #runs after finished threaded function
        self.thread[0].finished.connect(self.on_finished1)
        '''

    '''
    prints finished thread in terminal

    passes obj


    returns nothing
    '''
    def on_finished1(self):
        #self.thread[0].wait()
        print("Igniter Thread is Finished, Bytes in waiting = " + str(arduino.in_waiting))

    '''
    test button function thats starts 2nd thread

    passes obj

    returns nothing
    '''
    def start_test(self):

        #allows global variables in class
        global comport
        global arduino
        global i 

        #opens valves on gui
        #starts test sequence thread
        #checks if both steps are on and and calibration has been done
        if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
            self.is_on1 = True
            self.is_on2 = True
            self.switch1.setText("VALVE1: OPEN")
            self.switch1.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")
            self.switch2.setText("VALVE2: OPEN")
            self.switch2.setStyleSheet("QPushButton { background-color: red; color: black; font-size: 20px; }")

            #apply safety measures on mcu side software
            arduino.write(b'!')
            global test_on
            test_on = True
            self.buttons[1].setText("***TEST PROCEDURE***")
            self.thread[1] = MyThread()
            self.thread[1].function_type = 2
            self.test_start = time.time()-self.time
            #run
            self.thread[1].start()
            self.thread[1].finished.connect(self.on_finished2)
        self.test_done = time.time()-self.time

    '''
    sends thread update to terminal

    passes nothing 

    returns nothing
    '''
    def on_finished2(self):
        print("Test Thread Finished")
        self.buttons[1].setText("TEST PROCEDURE: ARMED")

        self.is_on1 = False
        self.is_on2 = False
        self.switch1.setText("VALVE1: CLOSE")
        self.switch1.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")
        self.switch2.setText("VALVE2: CLOSE")
        self.switch2.setStyleSheet("QPushButton { background-color: green; color: black; font-size: 20px; }")
        self.test_done = time.time()-self.time

    '''
    sends signal to calibrate

    passes obj

    returns nothing 
    '''
    def start_calibration(self):

        #sends to mcu
        arduino.write(b'@')

        time.sleep(4.2)# wait 4 seconds for 5000 sample

        #test button logic this should be made into a function
        self.buttons[0].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
        self.test = 1
        if self.arm1 == 1 and self.arm2 == 1 and self.test == 1:
            self.buttons[1].setText("TEST PROCEDURE: ARMED")
            self.buttons[1].setStyleSheet("QPushButton { background-color: green; color: black; font-size: 18px; }")
        else:
            self.buttons[1].setText("TEST PROCEDURE: NOT ARMED")
            self.buttons[1].setStyleSheet("QPushButton { background-color: red; color: black; font-size: 18px; }")

    '''opens labview button

       passes obj

       returns nothing
    '''
    def runLabView_Windows(self):
        #this is function is operating system specific and pretty much goes into the command line and oepns the LabView app
        subprocess.run('cd C:\\Program Files (x86)\\National Instruments\\LabVIEW 2023 & dir & start LabVIEW', shell = True)
    
    '''
    sends signal to start sending data

    passes obj

    returns nothing

    '''
    def start_plot(self):
        #only works if it has been calibrated
        if(self.test == 1):
            #starts updating GUI
            self.starting_plot = 1
            
            #start value equals 1
            arduino.write(b'9(')

            #changes timeout to accomodate slower comm
            arduino.timeout = 1
            #time to update on both sides
            time.sleep(0.1)

            #start time for data analysis
            self.time = time.time()

    '''
    force close everything

    passes obj

    returns nothing
    '''
    def abort(self):

        #breaks out of test sequence
        global SPARKED
        SPARKED = True

        #closes both valve
        arduino.write(b'4')
        time.sleep(0.05)
        #close valve 2
        arduino.write(b'5')

        print("abort function called")

    '''
    main function that updates visual part of screen

    passes obj

    returns nothing
    '''
    def update_plots(self):
        global ar, sparked
        begin = time.time()
        bits1 = b''
        ar.append(begin)
        flags = ''
        
        
        if  arduino.in_waiting > 0:

            #reads looking for status register
            temp = arduino.read()
            #print(temp)
            if temp != '':
                temp = temp[0]
            else:
                temp = 0
                
            #check if data is sent and 1 byte and not added junk to serial message
            #10 is \n and \r is 13
            if temp != 0 and temp != 10 and bin(temp) != 13:
                #dereferences -> turns bits into integers
                #goes through byte and checks each flag
                if temp == 7:
                    flags = arduino.read()
                    flags = flags[0]
                    for i in range(8):
                        if( (flags & 1) == 1):
                            self.status[i] = 1
                        else:
                            self.status[i] = 0
                        flags = flags >> 1

            #make sure on start up there are is no 1 in 1 place for all incoming data
            #this is default status register in little endian
            #print(self.status)
            if(self.status == [1,0,0,0,0,0,0,0]):
                self.start_bit = 1
                    
            #reads bytes after falg register for data sync purposes
            if(self.start_bit == 1 and self.status[0] == 1 and temp == 7):
                bits1 = arduino.read(self.bytes_read)
                #print(bits1)
                 
        #if byte array is 24 byts not blank these mean the same so which one will I keep
        if(len(bits1) == self.bytes_read and bits1 != '' and self.starting_plot == 1):
            #offset for each sensor reading (2 bytes a reading and 4 sensors)
            rounds = [0,8,16]

            #offset for low and high bit iof 10 bit pressure
            offset  = [0,2,4,6]
            
            pressures = [0] * 12
            index = 0
            #print(bits1)
            clk = 0

            for i in range(4):
            #round is 3 offset is 4
                for j in range(3):
                    #gets lower bit from serial buff
                    pt_low = bits1[rounds[j]*1 + offset[i] ]

                    #gets higher bits from serial buff
                    pt_high = bits1[rounds[j]*1 + (offset[i]+1)]

                    #combines them for 10 bit value
                    pressures[index] = (pt_high << 8) | pt_low
                    index += 1

            arr = [0] * 3
            medians  = []
            
            #MVF this filtering sorts and chooses median of 3 for 
            #pressure this is intended to mitigate inductive spikes and dips and data missalginmen
            #calculates all 3 medians from pressure


            for i in range(1,len(pressures)+1):
                arr[(i) % 3] = pressures[i-1]
                if i % 3 == 0:
                    arr.sort()
                    medians.append(arr[1])
        
            upper_bound = 1001
            lower_bound = 0
            #only update plot if value is within range
            #1001 because analog read value 984 gives 1001 and 983 gives 999
            if(medians[0] <= upper_bound and  medians[0] > lower_bound and self.status[4] != 1):
                self.new_data1 = medians[0]
    
            #updates bar item
            self.bar_item[0].setOpts(x=self.domain, height=self.new_data1)

            #checks pressure for ignition in seperate thread
            #changes tolerance to stop sparking
            
            combustion_pressure = 50
            if(self.new_data1 >= combustion_pressure):
                sparked = np.concatenate((sparked,[1]))
                SPARKED = True
            else:
                sparked = np.concatenate((sparked,[0]))

            
            #updates ducer 2 on gui
            if(medians[1] <= upper_bound and medians[1] > lower_bound and self.status[4] != 1):
                self.new_data2 = medians[1]
    
           
            self.bar_item[1].setOpts(x=self.domain, height=self.new_data2)

            #updates gui on ducer 3
            if( medians[2] <= upper_bound and medians[2] > lower_bound and self.status[4] != 1):
                self.new_data3 = medians[2]
    

            self.bar_item[2].setOpts(x=self.domain, height=self.new_data3)

            #updates gui on ducer 4
            if(medians[3] <= upper_bound and medians[3] > lower_bound and self.status[4] != 1):
                self.new_data4 = medians[3]
                
            self.bar_item[3].setOpts(x=self.domain, height=self.new_data4)
            
            #to speed up runtime of function I am writing all data
            
            #cuh = time.time()
            
            global s
            global avg
            global c
            global pt1, pt2, pt3, pt4, v1, v2, i1,a1,a2, a3, a4, x
            if( self.start_recording == 1):
               pt1 = np.concatenate((pt1,[self.new_data1]))
               pt2 = np.concatenate((pt2,[self.new_data2]))
               pt3 = np.concatenate((pt3,[self.new_data3]))
               pt4 = np.concatenate((pt4,[self.new_data4]))
               x   = np.concatenate((x,[(time.time() - self.time)]))

               v1  = np.concatenate((v1,[((self.status[6] or self.v1)*0.9)]))
               v2  = np.concatenate((v2,[((self.status[5] or self.v2)*0.8)]))
               i1  = np.concatenate((i1,[((self.status[4] or self.i1)*0.7)]))

               a1  = np.concatenate((a1,[(self.status[3]*0.9)]))
               a2  = np.concatenate((a2,[(self.status[2]*0.8)]))
               a3  = np.concatenate((a3,[(self.status[1]*0.7)]))
               a4  = np.concatenate((a4,[(self.status[7]*0.6)]))
    
               self.i += 1
               self.i1 = 0
            #print( (time.time()-cuh)*1000)
            '''

            if( self.start_recording == 1):
               txt_file.write(str(self.new_data1)+"\n"+str(self.new_data2)+"\n"+str(self.new_data3)+"\n"+str(self.new_data4)+"\n"+str( (self.status[6] or self.v1)*-100)+"\n")  
               txt_file.write(str( (self.status[5] or self.v2) *-90)+"\n"+str((self.status[4] or self.i1)*-80)+"\n"+str(time.time()- self.time)+"\n") #str(self.status[0]*-100
               txt_file.write(str(self.status[3]*-175)+"\n"+str(self.status[2]*-170)+"\n" +str(self.status[1]*-165)+"\n" + str(self.status[7]*-160)+"\n") 
               self.i += 1
            '''
        
                
        #checks flags and calls abort function
        '''
        if(self.status[7] == 1 or self.status[1] == 1 or self.status[2] == 1 or self.status[3] == 1):
            self.abort()
            #sends user abort  in terminal
            print("abort on GUI side")
        '''
        
        #computes refrsh rate of gui and logs time
        refresh_rate = str(math.ceil((time.time() - begin) * 1_000))
        curr_time = round( ((time.time() - self.time)),2)
        
        #seconds with 2 decimal places
        if curr_time > 100:
           curr_time *= round(pow(10,-9),2)

        #processes plot data on gui screen
        str_data1 = "  Psi = " + str(self.new_data1) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        str_data2 = "  Psi = " + str(self.new_data2) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        str_data3 = "  Psi = " + str(self.new_data3) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        str_data4 = "  Psi = " + str(self.new_data4) + "  Time = " + str(curr_time) + "s" + "  RR = " + refresh_rate + "ms"
        #print(str(self.new_data1))
        
        #updates plot data in GUI
        self.label1.setText(str_data1)
        self.label2.setText(str_data2)
        self.label3.setText(str_data3)
        self.label4.setText(str_data4)
        #print("process = " +  str((time.time()- t)*1_000))
        #print( (ar[len(ar)-1]+ar[len(ar)-2]))
        c += 1

#---------------------------------------------------------------------------------------
class FilterApp:
    def __init__(self, root, points):
        self.points = points
        self.root = root
        self.root.title("Filter Visualization")
        self.plot_data()
    
    def plot_data(self):
        global pt1, pt2, pt3, pt4, v1, v2, i1,a1,a2, a3, a4, window
        # Matplotlib Figure
        self.fig, (self.ax,self.bx,self.cx) = plt.subplots(3,1,figsize=(8, 6))
        self.fig.subplots_adjust(hspace=0.75)
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.root)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        self.ax.clear()
        
        avg = 0
        avg = x[len(x)-1]/len(pt1)

        # Plot sine wave
        self.ax.plot(x, pt1, label= "pt1")

        # Plot filtered function
        self.ax.plot(x, pt2, label= "pt2")

        # Plot sine wave
        self.ax.plot(x, pt3, label= "pt3")

        # Plot filtered function
        self.ax.plot(x, pt4, label= "pt4")

        self.bx.plot(x,v1, label= "valve 1")
        self.bx.plot(x,v2, label= "valve 2")
        self.bx.plot(x,i1,color = 'gray', label= "coil")

        self.cx.plot(x,a1, label= "pt1 abort")
        self.cx.plot(x,a2, label= "pt2 abort")
        self.cx.plot(x,a3, label= "pt3 abort")
        self.cx.plot(x,a4, label= "pt4 abort")

        self.ax.legend()
        self.ax.grid()
        self.ax.set_title("Engine Test Data @ RR = " + str(round( (avg*1_000),1)) + "ms")
        self.ax.set_xlabel("Time(s)")
        self.ax.set_ylabel("Pressure(Psi)")
        self.ax.set_ylim(0,1000)

        self.bx.legend()
        self.bx.grid()
        self.bx.set_title("Control")
        self.bx.set_xlabel("Time(s)")
        self.bx.set_ylabel("Activation")
        self.bx.set_ylim(-0.1,1)

        self.cx.legend()
        self.cx.grid()
        self.cx.set_title("Aborts")
        self.cx.set_xlabel("Time(s)")
        self.cx.set_ylabel("Abort Status")
        self.cx.set_ylim(-0.1,1)
        
        ulim = 700
        lowlim = 500
        low_vis = np.array([lowlim]*int(x.size))
        self.ax.plot(x,low_vis, linestyle = 'dotted', color = 'black')
        h_vis = np.array([ulim]*x.size)
        self.ax.plot(x,h_vis, linestyle = 'dotted', color = 'black')

        low_vis = np.array([100]*int(x.size))
        self.ax.plot(x,low_vis, linestyle = 'dotted',color = 'black')
        h_vis = np.array([250]*x.size)
        self.ax.plot(x,h_vis, linestyle = 'dotted',color = 'black')

       
        if test_on == True:
            start = np.linspace(0,1000,1000)
            start1 = np.linspace(0,1,1000)
            start_x = np.array([window.test_start]*1000)
            self.ax.plot(start_x,start, linestyle = 'dotted',color='black')
            self.bx.plot(start_x,start1, linestyle = 'dotted',color='black')
            self.cx.plot(start_x,start1, linestyle = 'dotted',color='black')

            end = np.linspace(0,1000,1000)
            end1 = np.linspace(0,1,1000)
            end_x = np.array([window.test_done]*1000)
            self.ax.plot(end_x,end, linestyle = 'dotted',color='black')
            self.bx.plot(end_x,end1, linestyle = 'dotted',color='black')
            self.cx.plot(end_x,end1, linestyle = 'dotted',color='black')

        # Draw the plot
        self.canvas.draw()

    
#---------------------Main Function-----------------------------------------
if __name__ == '__main__':
    test_on = False
    #variables for plotting
    points = 10_000
    pt1 = np.array([0])
    pt2 = np.array([0])
    pt3 = np.array([0])
    pt4 = np.array([0])

    v1 = np.array([0])
    v2 = np.array([0])
    i1 = np.array([0])

    a1 = np.array([0])
    a2 = np.array([0])
    a3 = np.array([0])
    a4 = np.array([0])

    sparked = np.array([0])

    x = np.array([0])
    #------------------------------------------

    ar = [0,0]
    s = 0
    avg = 0
    c = 1
    #B.G. = before GUI
    '''
    this want to this in main to make more sense
    '''
    #BEGINNING OF PROCESSING ( WHEN PROGRAM IS RAN)
    #get com port // differs by cable
    comport = input("Enter COMPORT:")

    #Initialize Serial Communication with Arduino
    #checks serial connection
    arduino=serial.Serial() 
    arduino.baudrate=115200#comm speed // has to be synced with arduino
    arduino.port= comport
    arduino.bytesize=serial.EIGHTBITS
    #used when not reading data // changes when reading function
    arduino.timeout = 0.01
    arduino.write_timeout = 0 #turns off write time out exception

    '''
    file set up for recording infrastructure
    '''
    #opening excel file premature for less wait time in the future
    dataframe = openpyxl.load_workbook("Recording_File.xlsx")
    sheet = dataframe.active
    sheet.delete_rows(2,sheet.max_row)
    dataframe.save("C:\\Users\\izuka\\Documents\\A_New_Ard_Proj\\Recording_File.xlsx")
    dataframe = openpyxl.load_workbook("Recording_File.xlsx")
    sheet = dataframe.active

    #opening texts file for real-time data collection in update function
    txt_file = open('data.txt', 'w+')
    #clears data in text file each app run
    txt_file.write("")
    txt_file.close()
    txt_file = open("data.txt", 'w')

    #checks if files are open and updates user in terminal
    if(txt_file):
        print("1: Raw File Ready 1/3")
    else:
        print("1: Raw File Not Ready")
    if(dataframe):
        print("2: Record File Available 2/3")
    else:
        print("2: Record File Not Available")


    #clears port buffer before anything is done
    if arduino.is_open:
        arduino.close()
    #trys to open com port if fails then trus 3 more times before quitting app
    try:
        arduino.open()
    except:
        i = 0
        while True and i < 3:
            arduino.port = comport
            time.sleep(0.5)
            arduino = serial.Serial()
            if arduino.is_open:
                break
            i += 1
        if( i == 3):
            print("Serial Connection Failed")
            sys.exit()

    #updates user in terminal
    if arduino.is_open:
        print("3: Connected to Arduino 3/3")
    else:
        print("3: Not Connected to Arduino")

    #NEEDED delay for arduino to set up
    time.sleep(2)
    app = QApplication(sys.argv)
    window = LFRE_GUI_Control_App()
    #password to start gui
    start_sequence = input("Enter Start Code:")
    while True:
        if(start_sequence == 'start'):
            break
        else:
            start_sequence = input("Enter Start Code:")

    #Application set up
    arduino.reset_input_buffer()
    window.show()
    app.exec_()

    #A.G = after GUI

    '''
    #writes array data to excel file
    for k in range( int((window.i/12)) ):
        for j in range(12):
            rec_data1 = sheet.cell(row = k+2, column = j+1)
            rec_data1.value = str("=" + window.recording_data[count]) 
            count += 1
    '''
    #saves data
    dataframe.save("C:\\Users\\izuka\\Documents\\A_New_Ard_Proj\\Recording_File.xlsx")
    if(window.start_recording == 1):
        print("Data Is Saved")
    '''
    if window.start_recording == 1:
        subprocess.run('C:\\Users\\izuka\Documents\\A_New_Ard_Proj\\Recording_File.xlsx', shell = True)
    '''
    if window.start_recording == 1:
        root_2 = tk.Tk()
        app_2 = FilterApp(root_2,window.i)
        root_2.mainloop()
    #clear all buffers when something messes up
    arduino.reset_output_buffer()
    #close serial port
    arduino.close()
    txt_file.close()
    #recording data from text file into excel
    txt_file = open("data.txt",'w')
    print( str((pt1.size*12*4) * 10**-3) + "kB's of data")
    for i in range(pt1.size):
        txt_file.write(str(pt1[i])+str(pt2[i])+str(pt3[i])+str(pt4[i]))
        txt_file.write(str(v1[i])+str(v2[i])+str(i1[i])+str(a1[i])+str(a2[i])) 
        txt_file.write(str(a3[i])+str(a4[i])+str(sparked[i]))
        txt_file.write('\n')
    count = 0
    txt_file.close()
    #exit program 
    sys.exit()

